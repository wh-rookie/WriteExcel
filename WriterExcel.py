from openpyxl import load_workbook


# 读取Excel文件中的数据position代表文件的位置
def ReadExcel(position, sheet_name="Sheet1"):
    """
    position: excel文件的位置
    sheet_name： 需要读取sheet表的名称，默认为第一个
    """
    # 定义位置
    excel = load_workbook(position)
    # 获取表名
    sheets = excel.get_sheet_names()
    if sheet_name == 'Sheet1':
        table = excel.get_sheet_by_name(sheets[0])
    else:
        table = excel.get_sheet_by_name(sheet_name)
    # 获取最大行数
    rows = table.max_row
    print(rows)
    columns = []
    columnsdata = []
    # 获取Excel文件目标列数据
    for i in range(1,rows+1):
        cellvalue = table.cell(row=i, column=1).value
        columns.append(cellvalue)
    # 去除所有None数据
    for i in range(len(columns)):
        if columns[i]:
            columnsdata.append(columns[i])
    return columnsdata

# 向某一列循环写入数据
def WriteExcel(position, TargetCol, lis, sheet_name="Sheet1"):
    """
    position:文件的位置
    TargetCol：需要写入文件的列
    lis：需要写入文件的数据
    sheet_name：需要写入的表名
    """
    excel = load_workbook(position)
   # 获取表名
    sheets = excel.get_sheet_names()
    if sheet_name == 'Sheet1':
        table = excel.get_sheet_by_name(sheets[0])
    else:
        table = excel.get_sheet_by_name(sheet_name)
    # 循环写入文件数据
    for i in range(len(lis)):
        table.cell(row=i+2, column=TargetCol).value = lis[i]
    # 保存文件
    excel.save(position)
